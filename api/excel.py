from openpyxl import Workbook
import openpyxl
from api.tariffs.models import Tariff
from django.conf import settings, Path

PATH = Path(settings.BASE_DIR, "api/excel_snippets/")


# PATH = "excel_snippets/tariff_snippet.xlsx"


class TariffToExcel:
    @staticmethod
    def export(tariffs):
        wb_snippet = openpyxl.load_workbook(Path(PATH, "tariff_snippet.xlsx"))
        # wb_snippet = openpyxl.load_workbook(PATH)
        sheet_snippet = wb_snippet.active
        snippet = [i.value for i in list(sheet_snippet.rows)[2]]
        ru_snippet = ["ID", "Название", "Регион", "Город", "Активен?", None, None]

        wb = Workbook()
        sheet = wb["Sheet"]
        sheet.title = "Тарифы"

        for col, cell in enumerate(snippet, 1):
            sheet.cell(row=1, column=col).value = ru_snippet[col - 1]

        for row, tariff in enumerate(tariffs, 2):
            for col, cell in enumerate(snippet, 1):
                if cell and "$" in cell:
                    if cell in ("$city", "$region"):
                        attr = getattr(tariff.city, cell.replace("$", ""))
                    else:
                        attr = getattr(tariff, cell.replace("$", ""))

                    sheet.cell(row=row, column=col).value = attr

        for row in sheet.rows:
            print([cell.value for cell in row])

        wb.save("test.xlsx")


TariffToExcel.export(Tariff.objects.all())
if __name__ == "__main__":
    TariffToExcel.export(Tariff.objects.all())
